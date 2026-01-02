import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
from concurrent.futures import ThreadPoolExecutor
def highlight_differences(old_csv_path, new_csv_path, output_path='compared_output.xlsx'):
    old_csv = pd.read_csv(old_csv_path, encoding='utf-8-sig')
    new_csv = pd.read_csv(new_csv_path, encoding='utf-8-sig')
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    font = Font(name='宋体')
    old_csv.set_index(old_csv.columns[0], inplace=True)
    new_csv.set_index(new_csv.columns[0], inplace=True)
    combined_index = sorted(set(old_csv.index).union(new_csv.index))
    def process_row(index):
        row = [index] + [new_csv.at[index, col] if index in new_csv.index and col in new_csv.columns else None for col in new_csv.columns]
        return row
    with ThreadPoolExecutor() as executor:
        rows = list(executor.map(process_row, combined_index))
    headers = [new_csv.index.name] + list(new_csv.columns)
    data = pd.DataFrame(rows, columns=headers)
    data_rows = list(dataframe_to_rows(data, index=False, header=True))
    wb = Workbook()
    ws = wb.active
    total_cells = 0
    highlighted_cells = 0
    for r_idx, row in enumerate(data_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = font
            total_cells += 1
            if r_idx > 1:
                index = data.iloc[r_idx - 2, 0]
                if index in new_csv.index and index not in old_csv.index:
                    cell.fill = highlight_fill
                    highlighted_cells += 1
                elif c_idx > 1:
                    col = data.columns[c_idx - 1]
                    if index in old_csv.index and col in old_csv.columns:
                        old_value = old_csv.get(col, pd.Series()).get(index, None)
                        new_value = new_csv.get(col, pd.Series()).get(index, None)
                        if old_value != new_value and not (pd.isnull(old_value) and pd.isnull(new_value)):
                            cell.fill = highlight_fill
                            highlighted_cells += 1
    wb.save(output_path)
    highlight_ratio = highlighted_cells / total_cells * 100
    print(f"Highlighted cells: {highlighted_cells}")
    print(f"Total cells: {total_cells}")
    print(f"Highlight ratio: {highlight_ratio:.2f}%")
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python 52.py <old_csv_path> <new_csv_path> [output_path]")
    else:
        highlight_differences(*sys.argv[1:4])
